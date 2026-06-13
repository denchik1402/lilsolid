from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, Response, send_from_directory
from datetime import datetime
import os
import re
import shutil
import logging
from urllib.parse import urlsplit

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Создать config.py из config.example.py ТОЛЬКО если его нет.
# ВАЖНО: Никогда не перезаписывать существующий config.py — там токены и ключи!
_config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.py')
_config_example = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.example.py')
if not os.path.exists(_config_path) and os.path.exists(_config_example):
    shutil.copy(_config_example, _config_path)
    print("Создан config.py — при необходимости отредактируйте TELEGRAM_BOT_TOKEN и др.")
elif os.path.exists(_config_path):
    pass  # config.py уже есть — не трогаем

# TG — папка с Telegram-ботом (telegram_notify используется сайтом для уведомлений)
_tg_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'TG')
if os.path.isdir(_tg_path) and _tg_path not in __import__('sys').path:
    __import__('sys').path.insert(0, _tg_path)

from extensions import db
from flask_caching import Cache

app = Flask(__name__)
_cache_config = {'CACHE_TYPE': 'SimpleCache', 'CACHE_DEFAULT_TIMEOUT': 300}
try:
    import config
    _cache_config['CACHE_TYPE'] = getattr(config, 'CACHE_TYPE', 'SimpleCache')
    if getattr(config, 'CACHE_TYPE', None) == 'FileSystemCache':
        _cache_config['CACHE_DIR'] = getattr(config, 'CACHE_DIR', './cache')
except ImportError:
    pass
cache = Cache(app, config=_cache_config)
def _get_secret_key():
    """SECRET_KEY: из env (рекомендуется), иначе config, иначе случайный для dev."""
    k = os.environ.get('SECRET_KEY')
    if k:
        return k
    if os.path.exists(_config_path):
        try:
            import config
            k = getattr(config, 'SECRET_KEY', None)
            if k:
                return k
        except ImportError:
            pass
    # Для продакшена задайте SECRET_KEY в переменных окружения
    return os.urandom(24).hex()
app.config['SECRET_KEY'] = _get_secret_key()
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(os.path.dirname(os.path.abspath(__file__)), "shop.db")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'static/images/products'
app.config['BANNERS_FOLDER'] = 'static/images/banners'
app.config['BLOG_FOLDER'] = 'static/images/blog'
app.config['WTF_CSRF_ENABLED'] = True
app.config['WTF_CSRF_CHECK_DEFAULT'] = True
# Для HTTPS: сессия должна передаваться по Secure
try:
    import config
    if getattr(config, 'SITE_URL', '').startswith('https://'):
        app.config['SESSION_COOKIE_SECURE'] = True
        app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
except ImportError:
    pass

db.init_app(app)

@app.before_request
def enforce_canonical_site_url():
    """
    Принудительный 301 на канонический SITE_URL (https + нужный host).
    Нужен для устранения дублей http/https и www/non-www.
    """
    if request.method not in ('GET', 'HEAD'):
        return None
    site_url = None
    if os.path.exists(_config_path):
        try:
            import config
            site_url = getattr(config, 'SITE_URL', None)
        except ImportError:
            site_url = None
    if not site_url or not isinstance(site_url, str) or not site_url.startswith('http'):
        return None
    if '127.0.0.1' in site_url or 'localhost' in site_url:
        return None
    target = urlsplit(site_url.rstrip('/'))
    if not target.scheme or not target.netloc:
        return None
    req_scheme = request.headers.get('X-Forwarded-Proto', request.scheme).split(',')[0].strip().lower()
    req_host = request.host.lower()
    target_scheme = target.scheme.lower()
    target_host = target.netloc.lower()
    if req_scheme == target_scheme and req_host == target_host:
        return None
    path_with_qs = request.full_path if request.query_string else request.path
    if path_with_qs.endswith('?'):
        path_with_qs = path_with_qs[:-1]
    return redirect(f'{target_scheme}://{target_host}{path_with_qs}', code=301)

from flask_wtf.csrf import CSRFProtect

csrf = CSRFProtect(app)

from sqlalchemy import event
from sqlalchemy.engine import Engine
from sqlalchemy.orm import joinedload

@event.listens_for(Engine, "connect")
def _set_sqlite_pragma(dbapi_connection, connection_record):
    """Оптимизация SQLite: WAL, кэш, меньше блокировок."""
    if "sqlite" in str(type(dbapi_connection)):
        cursor = dbapi_connection.cursor()
        cursor.execute("PRAGMA journal_mode=WAL")
        cursor.execute("PRAGMA synchronous=NORMAL")
        cursor.execute("PRAGMA cache_size=-64000")  # 64 MB
        cursor.execute("PRAGMA temp_store=MEMORY")
        cursor.close()

from models import Product, Category, Review, Order, OrderItem, TelegramUser, BotSetting, PromoCode, Banner, HomeBlock, DeviceModel, BlogPost
from seo_utils import normalize_device_model_name, generate_device_model_seo, make_url_slug

_telegram_bot_url_cache = None

def _get_telegram_bot_url():
    """URL бота для плавающей кнопки (t.me/username). Берёт из config или API."""
    global _telegram_bot_url_cache
    if _telegram_bot_url_cache is not None:
        return _telegram_bot_url_cache
    username = os.environ.get('TELEGRAM_BOT_USERNAME') or ''
    if not username and os.path.exists(_config_path):
        try:
            import config
            username = getattr(config, 'TELEGRAM_BOT_USERNAME', None) or ''
        except ImportError:
            pass
    if not username and os.path.exists(_config_path):
        try:
            import config
            token = getattr(config, 'TELEGRAM_BOT_TOKEN', None)
            if token:
                import urllib.request
                req = urllib.request.urlopen(f'https://api.telegram.org/bot{token}/getMe', timeout=5)
                data = __import__('json').loads(req.read().decode())
                if data.get('ok') and data.get('result', {}).get('username'):
                    username = data['result']['username']
        except Exception:
            pass
    _telegram_bot_url_cache = f"https://t.me/{str(username).lstrip('@')}" if username else None
    return _telegram_bot_url_cache

@app.route('/health')
def health():
    """Health check для мониторинга и балансировщиков нагрузки"""
    return jsonify({"status": "ok"}), 200


@app.route('/version')
def version():
    """Проверка версии деплоя — data-build в HTML должен совпадать"""
    return jsonify({"build": "20240317-mobile"}), 200


@app.errorhandler(404)
def not_found_error(e):
    """Страница 404 — не найдена"""
    return render_template('errors/404.html'), 404


@app.errorhandler(500)
def internal_error(e):
    """Страница 500 — ошибка сервера"""
    logger.exception("500 Internal Server Error: %s", e)
    try:
        db.session.rollback()
    except Exception:
        pass
    return render_template('errors/500.html'), 500


@app.context_processor
def inject_telegram_bot():
    return {'telegram_bot_url': _get_telegram_bot_url()}

def _invalidate_cache(*keys):
    """Сброс кэша при изменениях в админке."""
    to_delete = list(keys) if keys else [
        'nav_categories', 'index_data', 'index_banners', 'sitemap_xml'
    ]
    for k in to_delete:
        cache.delete(k)

def _get_nav_categories():
    """Категории для меню (кэш 5 мин)."""
    return Category.query.order_by(Category.name).all()

@app.context_processor
def inject_nav_categories():
    """Категории для выпадающего меню «Каталог» в шапке."""
    categories = cache.get('nav_categories')
    if categories is None:
        categories = _get_nav_categories()
        cache.set('nav_categories', categories, timeout=300)
    return {'nav_categories': categories}

SITE_PHONE_DEFAULT = '+7 (993) 596-82-25'
SITE_ADDRESS_DEFAULT = 'Москва, Ленинградское шоссе, 16А'
SITE_CITY_DEFAULT = 'Москва'


def _get_site_setting(name, default=''):
    val = os.environ.get(name)
    if val:
        return val
    if os.path.exists(_config_path):
        try:
            import config
            val = getattr(config, name, None)
        except ImportError:
            val = None
        if val:
            return val
    return default


def _phone_to_tel(display_phone):
    """+7 (993) 596-82-25 -> tel:+79935968225"""
    digits = ''.join(c for c in (display_phone or '') if c.isdigit())
    if not digits:
        return ''
    if digits.startswith('8') and len(digits) == 11:
        digits = '7' + digits[1:]
    elif len(digits) == 10:
        digits = '7' + digits
    return '+' + digits


@app.context_processor
def inject_site_contacts():
    phone = _get_site_setting('SITE_PHONE', SITE_PHONE_DEFAULT)
    return {
        'site_phone': phone,
        'site_phone_tel': _phone_to_tel(phone),
        'site_address': _get_site_setting('SITE_ADDRESS', SITE_ADDRESS_DEFAULT),
        'site_city': _get_site_setting('SITE_CITY', SITE_CITY_DEFAULT),
    }


def _get_yandex_metrika_id():
    """ID счётчика Яндекс.Метрики (только цифры). Пусто — счётчик не подключается."""
    if hasattr(_get_yandex_metrika_id, '_cached'):
        return _get_yandex_metrika_id._cached
    raw = _get_site_setting('YANDEX_METRIKA_ID', '')
    if not raw:
        _get_yandex_metrika_id._cached = None
        return None
    digits = ''.join(c for c in str(raw).strip() if c.isdigit())
    _get_yandex_metrika_id._cached = int(digits) if digits else None
    return _get_yandex_metrika_id._cached


@app.context_processor
def inject_yandex_metrika():
    return {'yandex_metrika_id': _get_yandex_metrika_id()}


@app.context_processor
def inject_seo_defaults():
    """SEO дефолты: canonical/og без query-параметров и с корректным хостом."""
    try:
        base = _sitemap_base_url()
    except Exception:
        base = request.url_root.rstrip('/')
    path = request.path if request.path.startswith('/') else '/' + request.path
    clean_url = (base + path).rstrip('/') if path != '/' else base + '/'
    return {
        'seo_site_url': base,
        'seo_canonical_url': clean_url,
        'seo_og_url': clean_url,
        'seo_favicon_url': base + '/favicon.svg',
        'seo_favicon_png_url': base + '/favicon.png',
        'seo_favicon_ico_url': base + '/favicon.ico',
    }


@app.context_processor
def inject_cart():
    """Корзина для виджета «В корзину» (количество по product_id)"""
    cart = session.get('cart', {})
    cart_qty = {int(k): v for k, v in cart.items() if str(k).isdigit()}
    return {'session_cart': cart, 'cart_qty': cart_qty}


import bleach

ALLOWED_TAGS_HTML = ['p', 'br', 'strong', 'em', 'u', 'b', 'i', 'a', 'ul', 'ol', 'li', 'h2', 'h3', 'h4', 'hr', 'span', 'div', 'blockquote']
ALLOWED_ATTRS_HTML = {'a': ['href', 'title', 'target', 'rel'], 'span': ['class'], 'div': ['class']}
ALLOWED_TAGS_BLOG = ALLOWED_TAGS_HTML + ['table', 'thead', 'tbody', 'tr', 'th', 'td', 'img']
ALLOWED_ATTRS_BLOG = {
    **ALLOWED_ATTRS_HTML,
    'img': ['src', 'alt', 'class', 'loading', 'width', 'height'],
    'table': ['class'],
    'th': ['class', 'colspan', 'rowspan'],
    'td': ['class', 'colspan', 'rowspan'],
}


def _sanitize_blog_html(value):
    if value is None or value == '':
        return ''
    return bleach.clean(str(value), tags=ALLOWED_TAGS_BLOG, attributes=ALLOWED_ATTRS_BLOG, strip=True)


def _wrap_blog_article_images(html: str) -> str:
    """Оборачивает inline-картинки для hover-zoom в тексте статьи."""
    if not html or '<img' not in html:
        return html
    html = re.sub(
        r'<span class="article-img-zoom">\s*(<img\b[^>]*>)\s*</span>',
        r'<div class="article-img-zoom">\1</div>',
        html,
        flags=re.IGNORECASE,
    )
    html = re.sub(
        r'<p(?:\s[^>]*)?>\s*(<img\b[^>]*>)\s*</p>',
        r'<div class="article-img-zoom">\1</div>',
        html,
        flags=re.IGNORECASE | re.DOTALL,
    )

    def _wrap_img(match: re.Match) -> str:
        start = match.start()
        before = html[max(0, start - 100):start]
        if re.search(r'class="article-img-zoom"[^>]*>\s*$', before):
            return match.group(0)
        return f'<div class="article-img-zoom">{match.group(0)}</div>'

    return re.sub(r'<img\b[^>]*>', _wrap_img, html, flags=re.IGNORECASE)


@app.template_filter('sanitize_html')
def sanitize_html(value):
    """Санитизация HTML для защиты от XSS (intro, description)."""
    if value is None or value == '':
        return ''
    return bleach.clean(str(value), tags=ALLOWED_TAGS_HTML, attributes=ALLOWED_ATTRS_HTML, strip=True)


@app.template_filter('sanitize_blog_html')
def sanitize_blog_html(value):
    if value is None or value == '':
        return ''
    return _sanitize_blog_html(value)


@app.template_filter('blog_article_images')
def blog_article_images(value):
    if value is None or value == '':
        return ''
    return _wrap_blog_article_images(str(value))


def _blog_cover_static_path(cover_image):
    if not cover_image:
        return None
    path = cover_image.strip().lstrip('/')
    if path.startswith('images/'):
        return path
    if path.startswith('products/'):
        return f'images/{path}'
    if path.startswith('blog/'):
        return f'images/{path}'
    return f'images/blog/{path}'


@app.template_filter('blog_cover_url')
def blog_cover_url(cover_image):
    static_path = _blog_cover_static_path(cover_image)
    if not static_path:
        return None
    return url_for('static', filename=static_path)


@app.template_filter('format_price')
def format_price(value):
    """Форматирует цену с пробелами как разделителями тысяч"""
    if value is None:
        return '0'
    return "{:,.0f}".format(float(value)).replace(",", " ")


@app.template_filter('product_alt')
def product_alt_filter(product):
    """Alt-текст для изображения товара (image_alt или name)"""
    if product is None:
        return ''
    return (product.image_alt or product.name or '').strip() or ''

@app.template_filter('product_url')
def product_url_filter(product):
    """URL страницы товара (ЧПУ)"""
    if product is None:
        return url_for('catalog')
    return url_for('product', product_slug=product.get_url_slug())


@app.route('/banner-click/<int:banner_id>')
def banner_click(banner_id):
    """Редирект по клику на баннер + учёт клика для A/B-статистики"""
    banner = Banner.query.filter_by(id=banner_id, is_active=True).first()
    if not banner:
        return redirect(url_for('index'))
    try:
        banner.clicks = (banner.clicks or 0) + 1
        db.session.commit()
    except Exception:
        db.session.rollback()
    if banner.product_id:
        p = db.session.get(Product, banner.product_id)
        target = url_for('product', product_slug=p.get_url_slug()) if p else url_for('catalog')
    else:
        target = banner.button_url or url_for('catalog')
    if target.startswith('http'):
        return redirect(target)
    return redirect(target if target.startswith('/') else '/' + target)


def _get_index_cached_data():
    """Данные главной (кэш 2 мин). Баннеры и A/B — отдельно, т.к. зависят от сессии."""
    data = cache.get('index_data')
    if data is not None:
        return data
    new_products = Product.query.order_by(Product.created_at.desc()).limit(8).all()
    popular_products = Product.query.order_by(Product.views.desc()).limit(4).all()
    promo_products = Product.query.filter(Product.old_price.isnot(None), Product.old_price > 0)\
        .order_by(Product.created_at.desc()).limit(6).all()
    try:
        hit_products = Product.query.filter(Product.is_hit == True).limit(8).all()
        if not hit_products:
            hit_products = Product.query.order_by(Product.views.desc()).limit(8).all()
    except Exception:
        hit_products = Product.query.order_by(Product.views.desc()).limit(8).all()
    recent_reviews = Review.query.filter_by(status='approved')\
        .order_by(Review.created_at.desc()).limit(10).all()
    home_blocks = HomeBlock.query.filter_by(is_active=True).order_by(HomeBlock.position).limit(3).all()
    data = {
        'new_products': new_products, 'popular_products': popular_products,
        'promo_products': promo_products, 'hit_products': hit_products,
        'recent_reviews': recent_reviews, 'home_blocks': home_blocks,
    }
    cache.set('index_data', data, timeout=120)
    return data

@app.route('/')
def index():
    """Главная страница"""
    data = _get_index_cached_data()
    new_products = data['new_products']
    popular_products = data['popular_products']
    promo_products = data['promo_products']
    hit_products = data['hit_products']
    recent_reviews = data['recent_reviews']
    home_blocks = data['home_blocks']
    
    # Слайды для карусели: баннеры (или товары, если баннеров нет). A/B: один вариант на группу.
    promo_slides_db = cache.get('index_banners')
    if promo_slides_db is None:
        promo_slides_db = Banner.query.filter_by(is_active=True).order_by(Banner.sort_order, Banner.id).limit(30).all()
        cache.set('index_banners', promo_slides_db, timeout=120)
    if not promo_slides_db:
        promo_slides_db = (promo_products or popular_products or new_products or Product.query.limit(5).all())[:5]
        promo_slides = list(promo_slides_db)
    else:
        # A/B-тест: баннеры с одинаковым ab_test_group — показываем один на сессию
        import random
        ab_groups = {}
        for b in promo_slides_db:
            g = b.ab_test_group or f'_b{b.id}'
            if g not in ab_groups:
                ab_groups[g] = []
            ab_groups[g].append(b)
        promo_slides = []
        for g, variants in ab_groups.items():
            if len(variants) == 1:
                chosen = variants[0]
            else:
                key = f'banner_ab_{g}'
                if key not in session:
                    session[key] = random.choice(range(len(variants)))
                chosen = variants[session[key]]
            promo_slides.append(chosen)
        promo_slides.sort(key=lambda x: (x.sort_order, x.id))
        promo_slides = promo_slides[:15]
        # Re-query Banner с joinedload(product) — объекты из кэша detached, lazy load .product падает
        banner_ids = [b.id for b in promo_slides if isinstance(b, Banner)]
        if banner_ids:
            fresh = {b.id: b for b in Banner.query.filter(Banner.id.in_(banner_ids)).options(joinedload(Banner.product)).all()}
            promo_slides = [fresh[b.id] if isinstance(b, Banner) else b for b in promo_slides]
        # Учитываем показы (impressions) — один commit вместо N
        try:
            for b in promo_slides:
                if isinstance(b, Banner):
                    b.impressions = (b.impressions or 0) + 1
            db.session.commit()
        except Exception:
            db.session.rollback()

    hit_product_ids = {p.id for p in (hit_products or [])}
    
    return render_template('index.html',
                         new_products=new_products,
                         popular_products=popular_products,
                         promo_products=promo_products,
                         hit_products=hit_products,
                         hit_product_ids=hit_product_ids,
                         recent_reviews=recent_reviews,
                         promo_slides=promo_slides,
                         home_blocks=home_blocks,
                         blog_posts=_query_published_blog_posts(limit=3))

@app.route('/catalog/<string:category_slug>')
@app.route('/catalog')
def catalog(category_slug=None):
    """Страница каталога с фильтрами (категория или модель устройства по ЧПУ-slug)."""
    category = None
    device_model_filter = None
    if category_slug:
        category = Category.query.filter_by(slug=category_slug).first()
        if not category:
            device_model_filter = DeviceModel.query.filter_by(slug=category_slug).first_or_404()

    model_q = request.args.get('model', '').strip()
    has_secondary_filters = any(
        request.args.get(k) for k in (
            'color', 'on_sale', 'exclusive', 'hit', 'in_stock', 'sort', 'view', 'price_min', 'price_max',
        )
    )
    if not category_slug and model_q and not has_secondary_filters and request.args.get('page', 1, type=int) == 1:
        dm_redirect = DeviceModel.query.filter(db.func.lower(DeviceModel.name) == model_q.lower()).first()
        if dm_redirect and dm_redirect.slug:
            extra = {k: v for k, v in request.args.items() if k not in ('model', 'page')}
            return redirect(url_for('catalog', category_slug=dm_redirect.slug, **extra), 301)

    if device_model_filter:
        q = Product.query.filter(db.func.lower(Product.model) == device_model_filter.name.lower())
    elif category:
        q = Product.query.filter_by(category_id=category.id)
    else:
        q = Product.query
    
    # Фильтр: только в наличии
    if request.args.get('in_stock') == '1':
        q = q.filter(Product.in_stock == True)
    
    # Фильтр: только со скидкой
    if request.args.get('on_sale') == '1':
        q = q.filter(Product.old_price.isnot(None), Product.old_price > 0)
    
    # Фильтр: модель (без учёта регистра)
    model_filter = device_model_filter.name if device_model_filter else request.args.get('model', '').strip()
    if model_filter and not device_model_filter:
        q = q.filter(db.func.lower(Product.model) == model_filter.lower())
    
    # Фильтр: цвет (без учёта регистра)
    color_filter = request.args.get('color', '').strip()
    if color_filter:
        q = q.filter(db.func.lower(Product.color) == color_filter.lower())
    
    # Фильтр: эксклюзив
    if request.args.get('exclusive') == '1':
        q = q.filter(Product.is_exclusive == True)
    
    # Фильтр: хит продаж
    if request.args.get('hit') == '1':
        q = q.filter(Product.is_hit == True)
    
    # Фильтр по цене (диапазон)
    price_min = request.args.get('price_min', type=float)
    price_max = request.args.get('price_max', type=float)
    if price_min is not None and price_min > 0:
        q = q.filter(Product.price >= price_min)
    if price_max is not None and price_max > 0:
        q = q.filter(Product.price <= price_max)
    
    # Сортировка
    sort = request.args.get('sort', 'default')
    if sort == 'price_asc':
        q = q.order_by(Product.price.asc())
    elif sort == 'price_desc':
        q = q.order_by(Product.price.desc())
    elif sort == 'newest':
        q = q.order_by(Product.created_at.desc())
    elif sort == 'popular':
        q = q.order_by(Product.views.desc())
    # default — без явной сортировки
    
    page = request.args.get('page', 1, type=int)
    per_page = min(max(request.args.get('per_page', 24, type=int), 12), 48)
    pagination = q.paginate(page=page, per_page=per_page, error_out=False)
    products = pagination.items
    categories = Category.query.all()
    
    # Диапазон цен для ползунка
    from sqlalchemy import func
    if category:
        price_stats = db.session.query(func.min(Product.price), func.max(Product.price)).filter(Product.category_id == category.id).first()
    else:
        price_stats = db.session.query(func.min(Product.price), func.max(Product.price)).first()
    price_range_min = int(price_stats[0] or 0)
    price_range_max = max(int(price_stats[1] or 10000), price_range_min + 100)
    
    view_mode = request.args.get('view', 'grid')
    if view_mode not in ('grid', 'list'):
        view_mode = 'grid'
    
    catalog_model_names = [m.name for m in _query_device_models()]
    filter_colors_fixed = ['Серый', 'Зеленый', 'Синий', 'Бежевый', 'Красный', 'Черный', 'Оранжевый', 'Фиолетовый', 'Желтый', 'Смешанный']

    base_q = Product.query.filter_by(category_id=category.id) if category else Product.query
    available_models = [r[0] for r in base_q.filter(
        Product.model.isnot(None), Product.model != ''
    ).with_entities(Product.model).distinct().all()]
    available_colors = [r[0] for r in base_q.filter(
        Product.color.isnot(None), Product.color != ''
    ).with_entities(Product.color).distinct().all()]
    if catalog_model_names:
        filter_models = [m for m in catalog_model_names if m in available_models]
        for m in available_models:
            if m not in filter_models:
                filter_models.append(m)
    else:
        filter_models = sorted(available_models)
    filter_colors = [c for c in filter_colors_fixed if c in available_colors]
    if model_filter and model_filter not in filter_models:
        filter_models = [model_filter] + filter_models
    if color_filter and color_filter not in filter_colors:
        filter_colors = [color_filter] + filter_colors

    if not device_model_filter and model_filter:
        device_model_filter = DeviceModel.query.filter(
            db.func.lower(DeviceModel.name) == model_filter.lower()
        ).first()

    filter_device_models = []
    for m in _query_device_models():
        if m.name in available_models or (model_filter and m.name.lower() == model_filter.lower()):
            filter_device_models.append(m)

    # Текущие значения фильтров для UI
    filters = {
        'in_stock': request.args.get('in_stock') == '1',
        'on_sale': request.args.get('on_sale') == '1',
        'price_min': price_min,
        'price_max': price_max,
        'sort': sort,
        'view': view_mode,
        'model': model_filter,
        'color': color_filter,
        'exclusive': request.args.get('exclusive') == '1',
        'hit': request.args.get('hit') == '1',
    }
    return render_template('catalog.html',
                         products=products,
                         pagination=pagination,
                         category=category,
                         categories=categories,
                         filters=filters,
                         filter_models=filter_models,
                         filter_colors=filter_colors,
                         price_range_min=price_range_min,
                         price_range_max=price_range_max,
                         device_model_filter=device_model_filter,
                         filter_device_models=filter_device_models)

@app.route('/product/<int:product_id>')
def product_redirect_id(product_id):
    """Редирект со старого URL /product/123 на ЧПУ /product/slug (301)"""
    product = db.session.get(Product, product_id)
    if not product:
        from flask import abort
        abort(404)
    return redirect(url_for('product', product_slug=product.get_url_slug()), code=301)


@app.route('/product/<product_slug>')
def product(product_slug):
    """Страница товара (ЧПУ: /product/iqos-iluma-i-one)"""
    product = Product.query.filter_by(slug=product_slug).first()
    if not product:
        if product_slug.isdigit():
            return redirect(url_for('product_redirect_id', product_id=int(product_slug)), code=301)
        from flask import abort
        abort(404)
    
    # Увеличиваем счетчик просмотров
    product.views += 1
    db.session.commit()
    
    # Получаем похожие товары (из той же категории)
    similar_products = Product.query.filter_by(category_id=product.category_id)\
                                   .filter(Product.id != product.id)\
                                   .limit(4).all()
    
    # Получаем одобренные отзывы на товар
    reviews = Review.query.filter_by(product_id=product.id, status='approved')\
                         .order_by(Review.created_at.desc()).all()
    
    from datetime import datetime, timedelta
    price_valid_until = (datetime.utcnow() + timedelta(days=365)).strftime('%Y-%m-%d')
    
    return render_template('product.html',
                         product=product,
                         similar_products=similar_products,
                         reviews=reviews,
                         price_valid_until=price_valid_until)

def _get_cart_products(session_cart):
    """Один запрос вместо N — загрузка всех товаров корзины."""
    if not session_cart:
        return [], 0
    ids = [int(k) for k in session_cart.keys() if str(k).isdigit()]
    if not ids:
        return [], 0
    products_map = {p.id: p for p in Product.query.filter(Product.id.in_(ids)).all()}
    items, total = [], 0
    for item_id, quantity in session_cart.items():
        try:
            pid = int(item_id)
        except (ValueError, TypeError):
            continue
        product = products_map.get(pid)
        if product:
            subtotal = product.price * quantity
            total += subtotal
            items.append({'product': product, 'quantity': quantity, 'subtotal': subtotal})
    return items, total


@app.route('/cart')
def cart():
    """Корзина"""
    session_cart = session.get('cart', {})
    cart_items, total = _get_cart_products(session_cart)
    cart_product_ids = [item['product'].id for item in cart_items]
    category_ids = [c for c in {item['product'].category_id for item in cart_items} if c is not None]
    similar_products = []
    if category_ids:
        q = Product.query.filter(Product.category_id.in_(category_ids))
        if cart_product_ids:
            q = q.filter(~Product.id.in_(cart_product_ids))
        similar_products = q.limit(8).all()
    if not similar_products:
        q = Product.query
        if cart_product_ids:
            q = q.filter(~Product.id.in_(cart_product_ids))
        similar_products = q.limit(8).all()
    return render_template('cart.html', cart_items=cart_items, total=total, similar_products=similar_products)

@csrf.exempt
@app.route('/add-to-cart/<int:product_id>', methods=['POST'])
def add_to_cart(product_id):
    """Добавление товара в корзину"""
    if 'cart' not in session:
        session['cart'] = {}
    try:
        quantity = max(1, min(99, int(request.form.get('quantity', 1))))
    except (ValueError, TypeError):
        quantity = 1
    
    if str(product_id) in session['cart']:
        session['cart'][str(product_id)] += quantity
    else:
        session['cart'][str(product_id)] = quantity
    
    session.modified = True
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True, 'cart_count': len(session['cart'])})
    
    return redirect(url_for('cart'))

@csrf.exempt
@app.route('/update-cart/<int:product_id>', methods=['POST'])
def update_cart(product_id):
    """Обновление количества товара в корзине"""
    if 'cart' in session:
        try:
            quantity = max(0, min(99, int(request.form.get('quantity', 0))))
        except (ValueError, TypeError):
            quantity = 0
        
        if quantity > 0:
            session['cart'][str(product_id)] = quantity
        else:
            session['cart'].pop(str(product_id), None)
        
        session.modified = True
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        items_data = []
        total = 0
        for item in _get_cart_products(session.get('cart', {}))[0]:
            p, qty, st = item['product'], item['quantity'], item['subtotal']
            items_data.append({'product_id': p.id, 'quantity': qty, 'subtotal': st, 'price': p.price})
            total += st
        return jsonify({
            'success': True,
            'cart_count': len(session['cart']),
            'total': total,
            'items': items_data
        })
    return redirect(url_for('cart'))

@app.route('/checkout', methods=['GET', 'POST'])
def checkout():
    """Оформление заказа"""
    cart_items, subtotal = _get_cart_products(session.get('cart', {}))
    discount, total = 0, subtotal
    promo_code = ''
    
    if request.method == 'POST':
        # Сумма считается только из корзины на сервере
        _, server_subtotal = _get_cart_products(session.get('cart', {}))
        if server_subtotal <= 0:
            return redirect(url_for('cart'))
        promo_code = (request.form.get('promo_code') or '').strip().upper()
        server_discount, server_total = 0, server_subtotal
        if promo_code:
            promo = PromoCode.query.filter_by(code=promo_code).first()
            if promo and promo.is_valid():
                server_discount, server_total = promo.apply_discount(server_subtotal)
                if server_discount > 0:
                    promo.used_count += 1

        order = Order(
            customer_name=request.form.get('name', '').strip()[:100],
            customer_phone=request.form.get('phone', '').strip()[:20],
            customer_email=request.form.get('email', '').strip()[:100],
            delivery_address=request.form.get('address', '').strip()[:500],
            delivery_method=request.form.get('delivery', 'pickup')[:50],
            payment_method=request.form.get('payment', 'store')[:50],
            comment=request.form.get('comment', '')[:500],
            total_amount=server_total,
            promo_code=promo_code if server_discount > 0 else None,
            discount_amount=server_discount
        )
        db.session.add(order)
        db.session.flush()

        for product_id, quantity in session.get('cart', {}).items():
            product = db.session.get(Product, int(product_id))
            if product and quantity > 0:
                db.session.add(OrderItem(
                    order_id=order.id,
                    product_id=product.id,
                    quantity=min(int(quantity), 99),
                    price=product.price
                ))

        db.session.commit()

        try:
            from telegram_notify import send_order_to_telegram
            ok, err = send_order_to_telegram(order)
            if not ok and err != "Telegram не настроен (TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID)":
                print(f"[Telegram] Ошибка: {err}")
        except Exception as e:
            print(f"[Telegram] Исключение: {e}")

        session['cart'] = {}
        session['last_order_id'] = order.id  # для проверки доступа к order_success
        session.modified = True

        return redirect(url_for('order_success', order_id=order.id))
    
    return render_template('checkout.html', cart_items=cart_items, total=total, subtotal=subtotal, discount=discount)

def _normalize_phone(s):
    """Нормализация телефона для сравнения"""
    if not s:
        return ''
    return str(s).replace(' ', '').replace('-', '').replace('(', '').replace(')', '').replace('+', '')


@app.route('/order-track', methods=['GET', 'POST'])
def order_track():
    """Отслеживание заказа по номеру и телефону"""
    order = None
    error = None
    if request.method == 'POST':
        order_number = (request.form.get('order_number') or '').strip().upper()
        phone = _normalize_phone((request.form.get('phone') or '').strip())
        if not order_number or not phone:
            error = 'Укажите номер заказа и телефон'
        else:
            order = Order.query.filter_by(order_number=order_number).first()
            if not order:
                error = 'Заказ не найден'
            elif _normalize_phone(order.customer_phone) != phone:
                error = 'Телефон не совпадает'
                order = None
    return render_template('order_track.html', order=order, error=error)


@app.route('/order-success/<int:order_id>')
def order_success(order_id):
    """Страница успешного заказа — доступ только после оформления этого заказа"""
    if session.get('last_order_id') != order_id:
        return redirect(url_for('index'))
    order = db.get_or_404(Order, order_id)
    session.pop('last_order_id', None)  # одноразовый доступ
    return render_template('order_success.html', order=order)

@app.route('/add-review', methods=['POST'])
def add_review():
    """Добавление отзыва (статус: на модерации). Только для покупателей."""
    product_id = request.form.get('product_id', type=int)
    if not product_id or not Product.query.get(product_id):
        flash('Неверный товар. Отзыв не сохранён.', 'danger')
        return redirect(url_for('index'))
    product = Product.query.get(product_id)
    email = (request.form.get('email') or '').strip().lower()
    phone = (request.form.get('phone') or '').strip().replace(' ', '').replace('-', '').replace('(', '').replace(')', '')

    if not email and not phone:
        flash('Укажите email или телефон для проверки покупки. Отзывы могут оставлять только покупатели.', 'danger')
        return redirect(url_for('product', product_slug=product.get_url_slug()))

    # Проверка: есть ли заказ с этим email/телефоном, содержащий данный товар
    from sqlalchemy import or_
    q = db.session.query(Order.id).join(OrderItem, Order.id == OrderItem.order_id).filter(
        OrderItem.product_id == product_id,
        Order.status != 'cancelled'
    )
    if email and phone:
        q = q.filter(or_(Order.customer_email.ilike(email), Order.customer_phone == phone))
    elif email:
        q = q.filter(Order.customer_email.ilike(email))
    else:
        q = q.filter(Order.customer_phone == phone)

    if not q.first():
        flash('Не найдена покупка этого товара по указанному email или телефону. Отзывы могут оставлять только покупатели.', 'danger')
        return redirect(url_for('product', product_slug=product.get_url_slug()))

    review = Review(
        product_id=product_id,
        customer_name=request.form['name'],
        rating=int(request.form['rating']),
        text=request.form['text'],
        status='pending'
    )
    db.session.add(review)
    db.session.commit()

    # Уведомление в Telegram о новом отзыве на модерации
    try:
        from telegram_notify import send_review_pending_to_telegram
        send_review_pending_to_telegram(review)
    except Exception as e:
        print(f"[Telegram] Отзыв: {e}")

    flash('Спасибо! Ваш отзыв отправлен на модерацию и будет опубликован после проверки.', 'success')
    return redirect(url_for('product', product_slug=product.get_url_slug()))

@app.route('/age-restricted')
def age_restricted():
    """Страница для пользователей младше 18 лет"""
    return render_template('age_restricted.html')

@app.route('/about')
def about():
    """О магазине"""
    return render_template('about.html')

@app.route('/delivery')
def delivery():
    """Доставка и оплата"""
    return render_template('delivery.html')

@app.route('/contacts')
def contacts():
    """Контакты"""
    return render_template('contacts.html')

@app.route('/faq')
def faq():
    """Часто задаваемые вопросы"""
    return render_template('faq.html')


def _query_published_blog_posts(limit=None):
    """Статьи блога; при отсутствии таблицы/колонок — пустой список."""
    try:
        q = BlogPost.query.filter_by(is_published=True).order_by(BlogPost.created_at.desc())
        if limit is not None:
            q = q.limit(limit)
        return q.all()
    except Exception as e:
        logger.warning('BlogPost query failed (run repair_schema.py): %s', e)
        try:
            db.session.rollback()
        except Exception:
            pass
        return []


def _get_blog_post_by_slug(slug):
    try:
        return BlogPost.query.filter_by(slug=slug, is_published=True).first()
    except Exception as e:
        logger.warning('BlogPost lookup failed: %s', e)
        try:
            db.session.rollback()
        except Exception:
            pass
        return None


@app.route('/blog')
def blog_index():
    """Блог и гайды"""
    posts = _query_published_blog_posts()
    return render_template('blog.html', posts=posts)


@app.route('/blog/<slug>')
def blog_post(slug):
    """Статья блога"""
    post = _get_blog_post_by_slug(slug)
    if not post:
        from flask import abort
        abort(404)
    others = [p for p in _query_published_blog_posts(limit=4) if p.id != post.id][:3]
    return render_template('blog_post.html', post=post, other_posts=others)


@app.route('/privacy')
def privacy():
    """Политика конфиденциальности и отказ от ответственности"""
    return render_template('privacy.html')

@app.route('/api/cart-count')
def cart_count():
    """API для получения количества товаров в корзине"""
    count = len(session.get('cart', {}))
    return jsonify({'count': count})


@csrf.exempt
@app.route('/api/clear-cart', methods=['POST'])
def clear_cart_api():
    """Очистка корзины"""
    session['cart'] = {}
    session.modified = True
    return jsonify({'success': True, 'count': 0})


@app.route('/api/search-suggestions')
def search_suggestions_api():
    """API для автодополнения при поиске"""
    from models import Product
    from search_utils import product_search_filter

    q = (request.args.get('q') or '').strip()
    if len(q) < 2:
        return jsonify({'suggestions': []})
    filt = product_search_filter(q, Product, in_stock_only=True)
    if filt is None:
        return jsonify({'suggestions': []})
    products = Product.query.filter(filt).order_by(Product.is_hit.desc(), Product.name).limit(10).all()
    suggestions = [
        {'name': p.name, 'url': url_for('product', product_slug=p.get_url_slug()), 'price': p.price}
        for p in products
    ]
    return jsonify({'suggestions': suggestions})


@app.route('/api/products-by-ids')
def products_by_ids_api():
    """API для блока «Вы смотрели» — товары по id из query ids=1,2,3"""
    ids_str = request.args.get('ids', '')
    ids = []
    for x in ids_str.replace(' ', '').split(','):
        try:
            i = int(x)
            if i > 0:
                ids.append(i)
        except ValueError:
            pass
    ids = list(dict.fromkeys(ids))[:12]  # уникальные, макс 12
    products = Product.query.filter(Product.id.in_(ids)).all() if ids else []
    products_by_id = {p.id: p for p in products}
    order = [products_by_id[i] for i in ids if i in products_by_id]
    items = []
    for p in order:
        img = p.all_images[0] if p.all_images else p.image
        items.append({
            'id': p.id,
            'name': p.name,
            'price': "{:,.0f}".format(float(p.price)).replace(",", " "),
            'old_price': "{:,.0f}".format(float(p.old_price)).replace(",", " ") if p.old_price else None,
            'image': img,
            'is_hit': bool(p.is_hit),
            'url': url_for('product', product_slug=p.get_url_slug()),
        })
    return jsonify({'products': items})


@app.route('/api/cart-items')
def cart_items_api():
    """API для боковой корзины — список товаров и сумма"""
    cart_items, total = _get_cart_products(session.get('cart', {}))
    items = [{'product_id': item['product'].id, 'slug': item['product'].get_url_slug(), 'name': item['product'].name,
              'quantity': item['quantity'], 'price': item['product'].price, 'subtotal': item['subtotal'],
              'url': url_for('product', product_slug=item['product'].get_url_slug()),
              'image': item['product'].image or (item['product'].all_images[0] if item['product'].all_images else None)}
             for item in cart_items]
    return jsonify({'items': items, 'total': total})

@csrf.exempt
@app.route('/api/validate-promo', methods=['POST'])
def validate_promo():
    """Проверка промокода и расчёт скидки"""
    code = (request.form.get('code') or (request.get_json(silent=True) or {}).get('code') or '').strip().upper()
    if not code:
        return jsonify({'success': False, 'error': 'Введите промокод'})
    promo = PromoCode.query.filter_by(code=code).first()
    if not promo:
        return jsonify({'success': False, 'error': 'Промокод не найден'})
    if not promo.is_valid():
        return jsonify({'success': False, 'error': 'Промокод недействителен или истёк'})
    _, total = _get_cart_products(session.get('cart', {}))
    discount, final = promo.apply_discount(total)
    if discount <= 0:
        return jsonify({'success': False, 'error': f'Минимальная сумма заказа для промокода: {promo.min_order:.0f} ₽'})
    return jsonify({
        'success': True,
        'discount': discount,
        'final': final,
        'message': f'Скидка {discount:.0f} ₽' + (f' ({promo.discount_value}%)' if promo.discount_type == 'percent' else '')
    })

@app.route('/admin')
def admin():
    """Админ-панель: вкладки Товары, Отзывы, Статистика"""
    login = _admin_or_login('admin.html')
    if login is not None:
        return login
    tab = request.args.get('tab', 'orders')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    orders = Order.query.order_by(Order.created_at.desc()).all()
    pending = Review.query.filter_by(status='pending').order_by(Review.created_at.desc()).all()
    products = Product.query.order_by(Product.name).all()
    categories = Category.query.all()
    promo_codes = PromoCode.query.order_by(PromoCode.created_at.desc()).all()
    pending_reviews_count = Review.query.filter_by(status='pending').count()
    products_count = Product.query.count()

    # Базовые запросы для статистики (без фильтра дат)
    order_base = Order.query
    review_base = Review.query

    # Фильтр по датам (по умолчанию — всё время)
    from datetime import datetime as dt, time
    if date_from:
        try:
            d_start = dt.combine(dt.strptime(date_from, '%Y-%m-%d').date(), time.min)
            order_base = order_base.filter(Order.created_at >= d_start)
            review_base = review_base.filter(Review.created_at >= d_start)
        except ValueError as e:
            logger.warning("Admin stats: invalid date_from %s: %s", date_from, e)
    if date_to:
        try:
            d_end = dt.combine(dt.strptime(date_to, '%Y-%m-%d').date(), time.max)
            order_base = order_base.filter(Order.created_at <= d_end)
            review_base = review_base.filter(Review.created_at <= d_end)
        except ValueError as e:
            logger.warning("Admin stats: invalid date_to %s: %s", date_to, e)

    # Статистика (с учётом фильтра)
    total_orders = order_base.count()
    new_orders = order_base.filter_by(status='new').count()
    completed_orders = order_base.filter_by(status='completed').count()
    total_revenue = order_base.filter(Order.status != 'cancelled').with_entities(db.func.sum(Order.total_amount)).scalar() or 0
    approved_reviews_count = review_base.filter_by(status='approved').count()

    # Данные для графиков
    orders_for_chart = order_base.filter(Order.status != 'cancelled').all()
    revenue_by_date = {}
    for o in orders_for_chart:
        d = o.created_at.date() if o.created_at else None
        if d:
            revenue_by_date[str(d)] = revenue_by_date.get(str(d), 0) + o.total_amount
    chart_revenue_labels = sorted(revenue_by_date.keys())
    chart_revenue_data = [revenue_by_date[k] for k in chart_revenue_labels]

    orders_all = order_base.all()
    status_counts = {'new': 0, 'processing': 0, 'completed': 0, 'cancelled': 0}
    for o in orders_all:
        status_counts[o.status] = status_counts.get(o.status, 0) + 1
    chart_status_labels = ['Новые', 'В работе', 'Выполнено', 'Отменено']
    chart_status_data = [status_counts['new'], status_counts['processing'], status_counts['completed'], status_counts['cancelled']]
    chart_status_colors = ['#ffc107', '#0d6efd', '#198754', '#dc3545']

    top_products_by_views = Product.query.order_by(Product.views.desc()).limit(10).all()
    banner_stats = Banner.query.order_by(Banner.sort_order, Banner.id).limit(15).all()

    banners = Banner.query.order_by(Banner.sort_order, Banner.id).all()
    home_blocks = HomeBlock.query.order_by(HomeBlock.position).all()
    products_for_link = Product.query.order_by(Product.name).limit(200).all()
    blog_posts = BlogPost.query.order_by(BlogPost.created_at.desc()).all()

    # Что сейчас в карусели на главной (для отображения в админке)
    class CarouselSlideInfo:
        """Слайд для отображения в админке (баннер или товар)"""
        __slots__ = ('type', 'title', 'obj', 'sort_order')
        def __init__(self, type_, title, obj=None, sort_order=None):
            self.type = type_
            self.title = title
            self.obj = obj
            self.sort_order = sort_order

    banners_in_carousel = Banner.query.filter_by(is_active=True).order_by(Banner.sort_order, Banner.id).limit(15).all()
    if banners_in_carousel:
        carousel_slides = [CarouselSlideInfo('banner', b.title, b, b.sort_order) for b in banners_in_carousel]
    else:
        promo_prods = Product.query.filter(Product.old_price.isnot(None), Product.old_price > 0).order_by(Product.created_at.desc()).limit(6).all()
        if not promo_prods:
            promo_prods = Product.query.order_by(Product.views.desc()).limit(5).all()
        if not promo_prods:
            promo_prods = Product.query.order_by(Product.created_at.desc()).limit(5).all()
        carousel_slides = [CarouselSlideInfo('product', p.name, p, i) for i, p in enumerate(promo_prods[:5])]

    return render_template('admin.html',
        orders=orders, reviews=pending, products=products, categories=categories, promo_codes=promo_codes, tab=tab,
        date_from=date_from, date_to=date_to,
        total_orders=total_orders, new_orders=new_orders, completed_orders=completed_orders,
        total_revenue=total_revenue, pending_reviews_count=pending_reviews_count,
        approved_reviews_count=approved_reviews_count, products_count=products_count,
        chart_revenue_labels=chart_revenue_labels, chart_revenue_data=chart_revenue_data,
        chart_status_labels=chart_status_labels, chart_status_data=chart_status_data, chart_status_colors=chart_status_colors,
        top_products_by_views=top_products_by_views, banner_stats=banner_stats,
        banners=banners, home_blocks=home_blocks, products_for_link=products_for_link,
        carousel_slides=carousel_slides,
        device_models=_query_device_models(),
        device_model_counts=_device_model_product_counts(),
        blog_posts=blog_posts,
        admin_role=_get_admin_role(),
        admin_can_delete_orders=_admin_can_delete_orders(),
    )


ADMIN_ROLE_ADMIN = 'admin'
ADMIN_ROLE_BOSS = 'boss'


def _get_config_secret(attr_name, env_name=None):
    env_name = env_name or attr_name
    val = os.environ.get(env_name)
    if not val and os.path.exists(_config_path):
        try:
            import config
            val = getattr(config, attr_name, None)
        except ImportError:
            pass
    return val


def _get_admin_secret():
    return _get_config_secret('ADMIN_SECRET')


def _get_boss_secret():
    return _get_config_secret('BOSS_SECRET')


def _set_admin_session(role):
    session['admin_logged_in'] = True
    session['admin_role'] = role
    session.permanent = True


def _get_admin_role():
    if not session.get('admin_logged_in'):
        return None
    return session.get('admin_role', ADMIN_ROLE_ADMIN)


def _admin_can_delete_orders():
    return _get_admin_role() == ADMIN_ROLE_ADMIN


def _authenticate_admin_key(key):
    """Проверка ключа: admin или boss. Возвращает роль или None."""
    key = (key or '').strip()
    if not key:
        return None
    admin_secret = _get_admin_secret()
    boss_secret = _get_boss_secret()
    if admin_secret and key == admin_secret:
        return ADMIN_ROLE_ADMIN
    if boss_secret and key == boss_secret:
        return ADMIN_ROLE_BOSS
    return None


def _admin_key_valid():
    """Проверка доступа в админку: сессия или ключ в URL (не рекомендуется)."""
    if session.get('admin_logged_in'):
        return True
    key = request.args.get('key')
    if key:
        role = _authenticate_admin_key(key)
        if role:
            _set_admin_session(role)
            return True
    return False


def _admin_or_login(template, **kwargs):
    """Если не авторизован — форма ввода ключа, иначе None"""
    if _admin_key_valid():
        return None
    error = None
    if request.method == 'POST' and request.form.get('key'):
        error = 'Неверный ключ'
    return render_template('admin_login.html', next_url=request.path, error=error)


def _query_device_models():
    return DeviceModel.query.order_by(DeviceModel.sort_order, DeviceModel.name).all()


def _device_model_product_counts():
    rows = db.session.query(Product.model, db.func.count(Product.id)).filter(
        Product.model.isnot(None), Product.model != ''
    ).group_by(Product.model).all()
    return {name: count for name, count in rows}


def _count_products_for_device_model(name):
    if not name:
        return 0
    return Product.query.filter(db.func.lower(Product.model) == name.lower()).count()


def _next_device_model_sort_order():
    max_order = db.session.query(db.func.max(DeviceModel.sort_order)).scalar()
    return (max_order or 0) + 10


def _save_device_model_seo(device_model, form):
    for field in ('image_alt', 'meta_description', 'meta_keywords'):
        val = (form.get(field) or '').strip()
        setattr(device_model, field, val or None)
    generated = generate_device_model_seo(device_model)
    for field in ('image_alt', 'meta_description', 'meta_keywords'):
        if not getattr(device_model, field):
            setattr(device_model, field, generated[field])


@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """Вход в админку по ключу (POST). Ключ не передаётся в URL."""
    next_url = request.form.get('next_url') or request.args.get('next') or url_for('admin')
    if request.method == 'POST':
        key = request.form.get('key')
        role = _authenticate_admin_key(key)
        if role:
            _set_admin_session(role)
            return redirect(next_url)
        return render_template('admin_login.html', next_url=next_url, error='Неверный ключ')
    return render_template('admin_login.html', next_url=next_url)


@app.route('/admin/logout')
def admin_logout():
    """Выход из админки"""
    session.pop('admin_logged_in', None)
    session.pop('admin_role', None)
    return redirect(url_for('admin_login'))


@app.route('/admin/products')
def admin_products():
    """Редирект на админку с вкладкой Товары"""
    if not _admin_key_valid():
        return render_template('admin_login.html', next_url=request.path, error=request.args.get('key') and 'Неверный ключ')
    return redirect(url_for('admin', tab='products'))


def _save_uploaded_image(file, product_id):
    """Сохраняет загруженный файл, возвращает имя файла или None"""
    if not file or not file.filename:
        return None
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else 'jpg'
    if ext not in ('jpg', 'jpeg', 'png', 'gif', 'webp'):
        return None
    filename = f"product_{product_id}_{abs(hash(file.filename)) % 100000}.{ext}"
    upload_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    file.save(upload_path)
    return filename


def _save_banner_image(file, prefix='banner'):
    """Сохраняет изображение баннера/блока, возвращает имя файла или None"""
    if not file or not file.filename:
        return None
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else 'jpg'
    if ext not in ('jpg', 'jpeg', 'png', 'gif', 'webp'):
        return None
    import time
    filename = f"{prefix}_{int(time.time())}_{abs(hash(file.filename)) % 10000}.{ext}"
    folder = app.config.get('BANNERS_FOLDER', 'static/images/banners')
    upload_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), folder, filename)
    os.makedirs(os.path.dirname(upload_path), exist_ok=True)
    file.save(upload_path)
    return filename


def _save_blog_image(file, prefix='blog'):
    """Сохраняет изображение блога, возвращает относительный путь blog/... или None."""
    if not file or not file.filename:
        return None
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else 'jpg'
    if ext not in ('jpg', 'jpeg', 'png', 'gif', 'webp'):
        return None
    import time
    sub = 'covers' if prefix == 'cover' else 'posts'
    filename = f"{prefix}_{int(time.time())}_{abs(hash(file.filename)) % 10000}.{ext}"
    folder = app.config.get('BLOG_FOLDER', 'static/images/blog')
    rel_path = f'blog/{sub}/{filename}'
    upload_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), folder, sub, filename)
    os.makedirs(os.path.dirname(upload_path), exist_ok=True)
    file.save(upload_path)
    return rel_path


def _unique_blog_slug(base_slug, exclude_id=None):
    slug = make_url_slug(base_slug)[:120] or 'post'
    candidate = slug
    n = 2
    while True:
        q = BlogPost.query.filter_by(slug=candidate)
        if exclude_id:
            q = q.filter(BlogPost.id != exclude_id)
        if not q.first():
            return candidate
        candidate = f'{slug}-{n}'[:120]
        n += 1


def _apply_blog_post_form(post, form, files):
    title = form.get('title', '').strip()
    if not title:
        raise ValueError('Заголовок обязателен')
    post.title = title
    slug_raw = form.get('slug', '').strip()
    post.slug = _unique_blog_slug(slug_raw or title, exclude_id=post.id if post.id else None)
    post.excerpt = form.get('excerpt', '').strip() or None
    content = form.get('content', '').strip()
    if not content:
        raise ValueError('Текст статьи обязателен')
    post.content = _sanitize_blog_html(content)
    post.meta_description = form.get('meta_description', '').strip() or None
    post.meta_keywords = form.get('meta_keywords', '').strip() or None
    post.cover_icon = form.get('cover_icon', '').strip() or 'fa-book-open'
    post.reading_minutes = max(1, int(form.get('reading_minutes', 5) or 5))
    post.is_published = form.get('is_published') == '1'
    cover_path = form.get('cover_image_path', '').strip()
    if cover_path:
        post.cover_image = cover_path.lstrip('/')
    cover_file = files.get('cover_image')
    if cover_file and cover_file.filename:
        saved = _save_blog_image(cover_file, 'cover')
        if saved:
            post.cover_image = saved
    post.updated_at = datetime.utcnow()


@app.route('/admin/product/add', methods=['GET', 'POST'])
def admin_product_add():
    """Добавление товара"""
    login = _admin_or_login('admin_product_add.html')
    if login is not None:
        return login
    categories = Category.query.all()
    if request.method == 'POST':
        try:
            cat_id = int(request.form.get('category_id'))
            name = request.form.get('name', '').strip()
            price = float(request.form.get('price', 0))
            if name and cat_id:
                slug = name.lower().replace(' ', '-')[:200]
                existing = Product.query.filter_by(slug=slug).first()
                if existing:
                    slug = f"{slug}-{Product.query.count()}"
                cost_val = request.form.get('cost')
                cost = float(cost_val) if cost_val else None
                model_val = request.form.get('model') or None
                color_val = request.form.get('color') or None
                meta_desc = (request.form.get('meta_description') or '').strip() or None
                meta_kw = (request.form.get('meta_keywords') or '').strip() or None
                img_alt = (request.form.get('image_alt') or '').strip() or None
                product = Product(name=name, slug=slug, price=price, cost=cost, category_id=cat_id, model=model_val, color=color_val, meta_description=meta_desc, meta_keywords=meta_kw, image_alt=img_alt)
                db.session.add(product)
                db.session.flush()
                img_file = request.files.get('image')
                if img_file and img_file.filename:
                    new_img = _save_uploaded_image(img_file, product.id)
                    if new_img:
                        product.image = new_img
                db.session.commit()
                _invalidate_cache()
                return redirect(url_for('admin', tab='products'))
        except (ValueError, TypeError) as e:
            logger.warning("Admin product add: %s", e)
    return render_template('admin_product_add.html', categories=categories, device_models=_query_device_models())


@app.route('/admin/product/<int:product_id>/edit', methods=['GET', 'POST'])
def admin_product_edit(product_id):
    """Редактирование товара (цена, изображение и т.д.)"""
    login = _admin_or_login('admin_product_edit.html')
    if login is not None:
        return login
    product = db.get_or_404(Product, product_id)
    if request.method == 'POST':
        try:
            product.price = float(request.form.get('price', product.price))
            product.name = request.form.get('name', product.name) or product.name
            old_val = request.form.get('old_price')
            product.old_price = float(old_val) if old_val and str(old_val).strip() else None
            cost_val = request.form.get('cost')
            product.cost = float(cost_val) if cost_val else None
            product.model = request.form.get('model') or None
            product.color = request.form.get('color') or None
            product.is_exclusive = request.form.get('is_exclusive') == 'on'
            product.is_hit = request.form.get('is_hit') == 'on'
            meta_desc = (request.form.get('meta_description') or '').strip() or None
            meta_kw = (request.form.get('meta_keywords') or '').strip() or None
            img_alt = (request.form.get('image_alt') or '').strip() or None
            product.meta_description = meta_desc
            product.meta_keywords = meta_kw
            product.image_alt = img_alt
            img_file = request.files.get('image')
            if img_file and img_file.filename:
                new_img = _save_uploaded_image(img_file, product_id)
                if new_img:
                    product.image = new_img
            db.session.commit()
            _invalidate_cache()
            return redirect(url_for('admin', tab='products'))
        except ValueError as e:
            logger.warning("Admin product edit: %s", e)
    return render_template('admin_product_edit.html', product=product, device_models=_query_device_models())


@app.route('/admin/category/<int:category_id>/edit', methods=['GET', 'POST'])
def admin_category_edit(category_id):
    """Редактирование meta категории (SEO)"""
    login = _admin_or_login('admin_category_edit.html')
    if login is not None:
        return login
    category = db.get_or_404(Category, category_id)
    if request.method == 'POST':
        desc = (request.form.get('description') or '').strip() or None
        meta_desc = (request.form.get('meta_description') or '').strip() or None
        meta_kw = (request.form.get('meta_keywords') or '').strip() or None
        category.description = desc
        category.meta_description = meta_desc
        category.meta_keywords = meta_kw
        db.session.commit()
        _invalidate_cache()
        return redirect(url_for('admin', tab='categories'))
    return render_template('admin_category_edit.html', category=category)


@app.route('/admin/device-model/add', methods=['GET', 'POST'])
def admin_device_model_add():
    """Добавление модели устройства"""
    login = _admin_or_login('admin_device_model_add.html')
    if login is not None:
        return login
    if request.method == 'POST':
        name = normalize_device_model_name((request.form.get('name') or '').strip())
        if not name:
            flash('Укажите название модели.', 'danger')
            return render_template('admin_device_model_add.html', default_sort_order=_next_device_model_sort_order())
        existing = DeviceModel.query.filter(db.func.lower(DeviceModel.name) == name.lower()).first()
        if existing:
            flash(f'Модель «{existing.name}» уже существует.', 'danger')
            return render_template('admin_device_model_add.html', default_sort_order=_next_device_model_sort_order())
        try:
            sort_raw = (request.form.get('sort_order') or '').strip()
            sort_order = int(sort_raw) if sort_raw else _next_device_model_sort_order()
            device_model = DeviceModel(name=name, sort_order=sort_order)
            _save_device_model_seo(device_model, request.form)
            db.session.add(device_model)
            db.session.commit()
            flash(f'Модель «{name}» создана.', 'success')
            return redirect(url_for('admin', tab='models'))
        except Exception as e:
            db.session.rollback()
            logger.warning("Admin device model add: %s", e)
            flash('Не удалось создать модель.', 'danger')
    return render_template('admin_device_model_add.html', default_sort_order=_next_device_model_sort_order())


@app.route('/admin/device-model/<int:model_id>/edit', methods=['GET', 'POST'])
def admin_device_model_edit(model_id):
    """Редактирование модели устройства"""
    login = _admin_or_login('admin_device_model_edit.html')
    if login is not None:
        return login
    device_model = db.get_or_404(DeviceModel, model_id)
    product_count = _count_products_for_device_model(device_model.name)
    if request.method == 'POST':
        name = normalize_device_model_name((request.form.get('name') or '').strip())
        if not name:
            flash('Укажите название модели.', 'danger')
            return render_template('admin_device_model_edit.html', device_model=device_model, product_count=product_count)
        duplicate = DeviceModel.query.filter(
            db.func.lower(DeviceModel.name) == name.lower(),
            DeviceModel.id != device_model.id,
        ).first()
        if duplicate:
            flash(f'Модель «{duplicate.name}» уже существует.', 'danger')
            return render_template('admin_device_model_edit.html', device_model=device_model, product_count=product_count)
        try:
            old_name = device_model.name
            sort_raw = (request.form.get('sort_order') or '').strip()
            device_model.name = name
            device_model.sort_order = int(sort_raw) if sort_raw else device_model.sort_order or 0
            _save_device_model_seo(device_model, request.form)
            if old_name != name:
                Product.query.filter(db.func.lower(Product.model) == old_name.lower()).update(
                    {'model': name}, synchronize_session=False
                )
            db.session.commit()
            flash('Модель сохранена.', 'success')
            return redirect(url_for('admin', tab='models'))
        except Exception as e:
            db.session.rollback()
            logger.warning("Admin device model edit: %s", e)
            flash('Не удалось сохранить модель.', 'danger')
    return render_template('admin_device_model_edit.html', device_model=device_model, product_count=product_count)


@app.route('/admin/device-model/<int:model_id>/delete', methods=['POST'])
def admin_device_model_delete(model_id):
    """Удаление модели устройства (только если нет товаров с этой моделью)"""
    if not _admin_key_valid():
        return "Доступ запрещён", 403
    device_model = db.get_or_404(DeviceModel, model_id)
    count = _count_products_for_device_model(device_model.name)
    if count:
        flash(
            f'Нельзя удалить «{device_model.name}»: модель указана у {count} товар(ов). '
            'Сначала измените модель у товаров или удалите их.',
            'danger',
        )
        return redirect(url_for('admin', tab='models'))
    name = device_model.name
    db.session.delete(device_model)
    db.session.commit()
    flash(f'Модель «{name}» удалена.', 'success')
    return redirect(url_for('admin', tab='models'))


@app.route('/admin/product/<int:product_id>/remove-from-carousel', methods=['POST'])
def admin_product_remove_from_carousel(product_id):
    """Убрать товар из карусели на главной (снять «Хит продаж»)"""
    if not _admin_key_valid():
        return "Доступ запрещён", 403
    product = db.get_or_404(Product, product_id)
    product.is_hit = False
    db.session.commit()
    _invalidate_cache()
    return redirect(url_for('admin', tab='banners'))


@app.route('/admin/product/<int:product_id>/delete', methods=['POST'])
def admin_product_delete(product_id):
    """Удаление товара"""
    if not _admin_key_valid():
        return "Доступ запрещён", 403
    product = db.get_or_404(Product, product_id)
    db.session.delete(product)
    db.session.commit()
    _invalidate_cache()
    return redirect(url_for('admin', tab='products'))


@app.route('/admin/products/bulk', methods=['POST'])
def admin_products_bulk():
    """Массовые операции с товарами"""
    if not _admin_key_valid():
        return "Доступ запрещён", 403
    ids = request.form.getlist('product_ids', type=int)
    action = request.form.get('bulk_action', '').strip()
    if not ids or not action:
        return redirect(url_for('admin', tab='products'))
    products = Product.query.filter(Product.id.in_(ids)).all()
    if not products:
        return redirect(url_for('admin', tab='products'))
    try:
        if action == 'category':
            cat_id = request.form.get('bulk_category_id', type=int)
            if cat_id and Category.query.get(cat_id):
                for p in products:
                    p.category_id = cat_id
        elif action == 'price':
            price = request.form.get('bulk_price', type=float)
            if price is not None and price >= 0:
                for p in products:
                    p.price = price
        elif action == 'old_price':
            old = request.form.get('bulk_old_price', type=float)
            for p in products:
                p.old_price = old if old is not None and old >= 0 else None
        elif action == 'in_stock':
            for p in products:
                p.in_stock = True
        elif action == 'out_of_stock':
            for p in products:
                p.in_stock = False
        elif action == 'delete':
            for p in products:
                db.session.delete(p)
        else:
            return redirect(url_for('admin', tab='products'))
        db.session.commit()
        _invalidate_cache()
    except (ValueError, TypeError) as e:
        logger.warning("Admin bulk: %s", e)
    return redirect(url_for('admin', tab='products'))


@app.route('/admin/review/<int:review_id>/<action>')
def admin_review_action(review_id, action):
    """Одобрение/отклонение отзыва из админки"""
    if not _admin_key_valid():
        return "Доступ запрещён", 403
    review = db.get_or_404(Review, review_id)
    if action == 'approve':
        review.status = 'approved'
    elif action == 'reject':
        review.status = 'rejected'
    else:
        return "Неверное действие", 400
    db.session.commit()
    _invalidate_cache()
    return redirect(url_for('admin', tab='reviews'))


def _parse_promo_form():
    """Парсинг формы промокода, возвращает dict для PromoCode"""
    from datetime import datetime as dt
    code = (request.form.get('code') or '').strip().upper()
    discount_type = request.form.get('discount_type', 'percent')
    discount_value = float(request.form.get('discount_value', 0))
    min_order = float(request.form.get('min_order') or 0)
    max_uses_raw = request.form.get('max_uses', '').strip()
    max_uses = int(max_uses_raw) if max_uses_raw and max_uses_raw.isdigit() else None
    valid_from_raw = request.form.get('valid_from', '').strip()
    valid_from = dt.strptime(valid_from_raw, '%Y-%m-%dT%H:%M') if valid_from_raw else None
    valid_until_raw = request.form.get('valid_until', '').strip()
    valid_until = dt.strptime(valid_until_raw, '%Y-%m-%dT%H:%M') if valid_until_raw else None
    is_active = request.form.get('is_active') == '1'
    return {
        'code': code, 'discount_type': discount_type, 'discount_value': discount_value,
        'min_order': min_order, 'max_uses': max_uses,
        'valid_from': valid_from, 'valid_until': valid_until, 'is_active': is_active
    }


@app.route('/admin/promo/add', methods=['GET', 'POST'])
def admin_promo_add():
    """Добавление промокода (только админ)"""
    login = _admin_or_login('admin_promo_add.html')
    if login is not None:
        return login
    if request.method == 'POST':
        try:
            data = _parse_promo_form()
            if not data['code']:
                raise ValueError('Код обязателен')
            if PromoCode.query.filter_by(code=data['code']).first():
                return render_template('admin_promo_add.html', error='Промокод с таким кодом уже существует')
            promo = PromoCode(**data)
            db.session.add(promo)
            db.session.commit()
            return redirect(url_for('admin', tab='promo'))
        except (ValueError, TypeError) as e:
            logger.warning("Admin promo add: %s", e)
    return render_template('admin_promo_add.html')


@app.route('/admin/promo/<int:promo_id>/edit', methods=['GET', 'POST'])
def admin_promo_edit(promo_id):
    """Редактирование промокода"""
    login = _admin_or_login('admin_promo_edit.html')
    if login is not None:
        return login
    promo = db.get_or_404(PromoCode, promo_id)
    if request.method == 'POST':
        try:
            data = _parse_promo_form()
            if not data['code']:
                raise ValueError('Код обязателен')
            other = PromoCode.query.filter_by(code=data['code']).first()
            if other and other.id != promo_id:
                return render_template('admin_promo_edit.html', promo=promo, error='Промокод с таким кодом уже существует')
            for k, v in data.items():
                setattr(promo, k, v)
            db.session.commit()
            return redirect(url_for('admin', tab='promo'))
        except (ValueError, TypeError) as e:
            logger.warning("Admin promo edit: %s", e)
    return render_template('admin_promo_edit.html', promo=promo)


@app.route('/admin/order/<int:order_id>/delete', methods=['POST'])
def admin_order_delete(order_id):
    """Удаление заказа (только роль admin)"""
    if not _admin_key_valid():
        return "Доступ запрещён", 403
    if not _admin_can_delete_orders():
        flash('Удаление заказов доступно только администратору.', 'warning')
        return redirect(url_for('admin', tab='orders'))
    order = db.get_or_404(Order, order_id)
    db.session.delete(order)
    db.session.commit()
    return redirect(url_for('admin', tab='orders'))


@app.route('/admin/orders/export')
def admin_orders_export():
    """Экспорт заказов в Excel или CSV"""
    if not _admin_key_valid():
        return "Доступ запрещён", 403
    fmt = request.args.get('format', 'xlsx').lower()
    orders = Order.query.order_by(Order.created_at.desc()).all()
    if fmt == 'csv':
        import csv
        import io
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['№', 'Дата', 'Клиент', 'Телефон', 'Email', 'Адрес', 'Доставка', 'Оплата', 'Сумма', 'Скидка', 'Промокод', 'Статус', 'Комментарий'])
        for o in orders:
            writer.writerow([
                o.order_number, o.created_at.strftime('%d.%m.%Y %H:%M') if o.created_at else '',
                o.customer_name, o.customer_phone or '', o.customer_email or '',
                o.delivery_address or '', o.delivery_method or '', o.payment_method or '',
                o.total_amount, o.discount_amount or 0, o.promo_code or '',
                o.status, o.comment or ''
            ])
        output.seek(0)
        return Response(output.getvalue(), mimetype='text/csv; charset=utf-8-sig',
                       headers={'Content-Disposition': 'attachment; filename=orders.csv'})
    # Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        return "Установите openpyxl: pip install openpyxl", 500
    wb = Workbook()
    ws = wb.active
    ws.title = 'Заказы'
    headers = ['№', 'Дата', 'Клиент', 'Телефон', 'Email', 'Адрес', 'Доставка', 'Оплата', 'Сумма', 'Скидка', 'Промокод', 'Статус', 'Комментарий']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
    for row, o in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=o.order_number)
        ws.cell(row=row, column=2, value=o.created_at.strftime('%d.%m.%Y %H:%M') if o.created_at else '')
        ws.cell(row=row, column=3, value=o.customer_name)
        ws.cell(row=row, column=4, value=o.customer_phone or '')
        ws.cell(row=row, column=5, value=o.customer_email or '')
        ws.cell(row=row, column=6, value=o.delivery_address or '')
        ws.cell(row=row, column=7, value=o.delivery_method or '')
        ws.cell(row=row, column=8, value=o.payment_method or '')
        ws.cell(row=row, column=9, value=o.total_amount)
        ws.cell(row=row, column=10, value=o.discount_amount or 0)
        ws.cell(row=row, column=11, value=o.promo_code or '')
        ws.cell(row=row, column=12, value=o.status)
        ws.cell(row=row, column=13, value=o.comment or '')
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                   headers={'Content-Disposition': 'attachment; filename=orders.xlsx'})


@app.route('/admin/promo/<int:promo_id>/delete', methods=['POST'])
def admin_promo_delete(promo_id):
    """Удаление промокода"""
    if not _admin_key_valid():
        return "Доступ запрещён", 403
    promo = db.get_or_404(PromoCode, promo_id)
    db.session.delete(promo)
    db.session.commit()
    return redirect(url_for('admin', tab='promo'))


@app.route('/admin/blog/add', methods=['GET', 'POST'])
def admin_blog_add():
    """Добавление статьи блога"""
    login = _admin_or_login('admin_blog_edit.html')
    if login is not None:
        return login
    if request.method == 'POST':
        try:
            post = BlogPost(created_at=datetime.utcnow())
            _apply_blog_post_form(post, request.form, request.files)
            db.session.add(post)
            db.session.commit()
            _invalidate_cache('sitemap_xml')
            flash('Статья опубликована.' if post.is_published else 'Черновик сохранён.', 'success')
            return redirect(url_for('admin', tab='blog'))
        except (ValueError, TypeError) as e:
            flash(str(e), 'danger')
        except (PermissionError, OSError) as e:
            logger.exception("Admin blog add: %s", e)
            flash('Ошибка сохранения изображения. Проверьте права на static/images/blog.', 'danger')
    return render_template('admin_blog_edit.html', post=None)


@app.route('/admin/blog/<int:post_id>/edit', methods=['GET', 'POST'])
def admin_blog_edit(post_id):
    """Редактирование статьи блога"""
    login = _admin_or_login('admin_blog_edit.html')
    if login is not None:
        return login
    post = db.get_or_404(BlogPost, post_id)
    if request.method == 'POST':
        try:
            _apply_blog_post_form(post, request.form, request.files)
            db.session.commit()
            _invalidate_cache('sitemap_xml')
            flash('Статья сохранена.', 'success')
            return redirect(url_for('admin', tab='blog'))
        except (ValueError, TypeError) as e:
            flash(str(e), 'danger')
        except (PermissionError, OSError) as e:
            logger.exception("Admin blog edit: %s", e)
            flash('Ошибка сохранения изображения. Проверьте права на static/images/blog.', 'danger')
    return render_template('admin_blog_edit.html', post=post)


@app.route('/admin/blog/<int:post_id>/delete', methods=['POST'])
def admin_blog_delete(post_id):
    if not _admin_key_valid():
        return "Доступ запрещён", 403
    post = db.get_or_404(BlogPost, post_id)
    db.session.delete(post)
    db.session.commit()
    _invalidate_cache('sitemap_xml')
    flash('Статья удалена.', 'success')
    return redirect(url_for('admin', tab='blog'))


@app.route('/admin/blog/upload-image', methods=['POST'])
def admin_blog_upload_image():
    """Загрузка изображения для вставки в текст статьи."""
    if not _admin_key_valid():
        return jsonify({'error': 'Доступ запрещён'}), 403
    img_file = request.files.get('image')
    saved = _save_blog_image(img_file, 'post')
    if not saved:
        return jsonify({'error': 'Недопустимый формат (JPG, PNG, GIF, WebP)'}), 400
    return jsonify({'url': url_for('static', filename=f'images/{saved}')})


@app.route('/admin/banner/add', methods=['GET', 'POST'])
def admin_banner_add():
    """Добавление баннера карусели"""
    login = _admin_or_login('admin_banner_edit.html')
    if login is not None:
        return login
    if request.method == 'POST':
        try:
            title = request.form.get('title', '').strip()
            if not title:
                raise ValueError('Заголовок обязателен')
            img_file = request.files.get('image')
            if not img_file or not img_file.filename:
                raise ValueError('Загрузите изображение')
            filename = _save_banner_image(img_file, 'banner')
            if not filename:
                raise ValueError('Недопустимый формат изображения')
            banner = Banner(image=filename, title=title)
            banner.subtitle = request.form.get('subtitle', '').strip() or None
            banner.button_text = request.form.get('button_text', '').strip() or 'Купить сейчас'
            pid = request.form.get('product_id')
            banner.product_id = int(pid) if pid else None
            banner.button_url = request.form.get('button_url', '').strip() or None
            banner.badge_type = request.form.get('badge_type', '').strip() or None
            banner.sort_order = int(request.form.get('sort_order', 0) or 0)
            banner.ab_test_group = request.form.get('ab_test_group', '').strip() or None
            db.session.add(banner)
            db.session.commit()
            _invalidate_cache()
            return redirect(url_for('admin', tab='banners'))
        except (ValueError, TypeError) as e:
            logger.warning("Admin banner add: %s", e)
        except (PermissionError, OSError) as e:
            logger.exception("Admin banner add: ошибка записи файла: %s", e)
            flash('Ошибка сохранения изображения. Проверьте права на static/images/banners.', 'danger')
    return render_template('admin_banner_edit.html', banner=None, products_for_link=Product.query.order_by(Product.name).limit(200).all())


@app.route('/admin/banner/<int:banner_id>/edit', methods=['GET', 'POST'])
def admin_banner_edit(banner_id):
    """Редактирование баннера"""
    login = _admin_or_login('admin_banner_edit.html')
    if login is not None:
        return login
    banner = db.get_or_404(Banner, banner_id)
    if request.method == 'POST':
        try:
            banner.title = request.form.get('title', '').strip() or banner.title
            banner.subtitle = request.form.get('subtitle', '').strip() or None
            banner.button_text = request.form.get('button_text', '').strip() or 'Купить сейчас'
            pid = request.form.get('product_id')
            banner.product_id = int(pid) if pid else None
            banner.button_url = request.form.get('button_url', '').strip() or None
            banner.badge_type = request.form.get('badge_type', '').strip() or None
            banner.sort_order = int(request.form.get('sort_order', 0) or 0)
            banner.ab_test_group = request.form.get('ab_test_group', '').strip() or None
            img_file = request.files.get('image')
            if img_file and img_file.filename:
                new_img = _save_banner_image(img_file, 'banner')
                if new_img:
                    banner.image = new_img
            db.session.commit()
            _invalidate_cache()
            return redirect(url_for('admin', tab='banners'))
        except (ValueError, TypeError) as e:
            logger.warning("Admin banner edit: %s", e)
        except (PermissionError, OSError) as e:
            logger.exception("Admin banner edit: ошибка записи файла: %s", e)
            flash('Ошибка сохранения изображения. Проверьте права на static/images/banners.', 'danger')
    return render_template('admin_banner_edit.html', banner=banner, products_for_link=Product.query.order_by(Product.name).limit(200).all())


@app.route('/admin/banner/<int:banner_id>/set_order', methods=['POST'])
def admin_banner_set_order(banner_id):
    """Установить порядок баннера (число)"""
    if not _admin_key_valid():
        return "Доступ запрещён", 403
    banner = db.get_or_404(Banner, banner_id)
    try:
        order_val = int(request.form.get('sort_order', 0) or 0)
        banner.sort_order = order_val
        db.session.commit()
        _invalidate_cache()
    except (ValueError, TypeError):
        pass
    return redirect(url_for('admin', tab='banners'))


@app.route('/admin/banner/<int:banner_id>/move/<direction>')
def admin_banner_move(banner_id, direction):
    """Переместить баннер выше/ниже (изменить порядок)"""
    if not _admin_key_valid():
        return "Доступ запрещён", 403
    banner = db.get_or_404(Banner, banner_id)
    all_banners = Banner.query.order_by(Banner.sort_order, Banner.id).all()
    idx = next((i for i, b in enumerate(all_banners) if b.id == banner.id), None)
    if idx is None:
        return redirect(url_for('admin', tab='banners'))
    swap_idx = idx - 1 if direction == 'up' and idx > 0 else (idx + 1 if direction == 'down' and idx < len(all_banners) - 1 else None)
    if swap_idx is not None:
        b1, b2 = all_banners[idx], all_banners[swap_idx]
        b1.sort_order, b2.sort_order = b2.sort_order, b1.sort_order
        db.session.commit()
        _invalidate_cache()
    return redirect(url_for('admin', tab='banners'))


@app.route('/admin/banner/<int:banner_id>/delete', methods=['POST'])
def admin_banner_delete(banner_id):
    if not _admin_key_valid():
        return "Доступ запрещён", 403
    banner = db.get_or_404(Banner, banner_id)
    db.session.delete(banner)
    db.session.commit()
    _invalidate_cache()
    return redirect(url_for('admin', tab='banners'))


@app.route('/admin/homeblock/add', methods=['GET', 'POST'])
def admin_homeblock_add():
    """Добавление блока «Премиальные устройства»"""
    login = _admin_or_login('admin_homeblock_edit.html')
    if login is not None:
        return login
    if request.method == 'POST':
        try:
            position = int(request.form.get('position', 1) or 1)
            if position not in (1, 2, 3):
                position = 1
            title = request.form.get('title', '').strip()
            if not title:
                raise ValueError('Заголовок обязателен')
            img_file = request.files.get('image')
            if not img_file or not img_file.filename:
                raise ValueError('Загрузите изображение')
            filename = _save_banner_image(img_file, 'block')
            if not filename:
                raise ValueError('Недопустимый формат изображения')
            block = HomeBlock(position=position, image=filename, title=title)
            block.description = request.form.get('description', '').strip() or None
            block.button_text = request.form.get('button_text', '').strip() or None
            block.button_url = request.form.get('button_url', '').strip() or None
            db.session.add(block)
            db.session.commit()
            _invalidate_cache()
            return redirect(url_for('admin', tab='banners'))
        except (ValueError, TypeError) as e:
            logger.warning("Admin homeblock add: %s", e)
        except (PermissionError, OSError) as e:
            logger.exception("Admin homeblock add: ошибка записи файла: %s", e)
            flash('Ошибка сохранения изображения. Проверьте права на static/images/banners.', 'danger')
        except Exception as e:
            logger.exception("Admin homeblock add: неожиданная ошибка: %s", e)
            db.session.rollback()
            flash(f'Ошибка: {type(e).__name__}. Проверьте логи на сервере: journalctl -u lilstore -n 50', 'danger')
    return render_template('admin_homeblock_edit.html', block=None)


@app.route('/admin/homeblock/<int:block_id>/edit', methods=['GET', 'POST'])
def admin_homeblock_edit(block_id):
    """Редактирование блока «Премиальные устройства»"""
    login = _admin_or_login('admin_homeblock_edit.html')
    if login is not None:
        return login
    block = db.get_or_404(HomeBlock, block_id)
    if request.method == 'POST':
        try:
            block.title = request.form.get('title', '').strip() or block.title
            block.description = request.form.get('description', '').strip() or None
            block.button_text = request.form.get('button_text', '').strip() or None
            block.button_url = request.form.get('button_url', '').strip() or None
            block.position = int(request.form.get('position', 1) or 1)
            if block.position not in (1, 2, 3):
                block.position = 1
            img_file = request.files.get('image')
            if img_file and img_file.filename:
                new_img = _save_banner_image(img_file, 'block')
                if new_img:
                    block.image = new_img
            db.session.commit()
            _invalidate_cache()
            return redirect(url_for('admin', tab='banners'))
        except (ValueError, TypeError) as e:
            logger.warning("Admin homeblock edit: %s", e)
        except (PermissionError, OSError) as e:
            logger.exception("Admin homeblock edit: ошибка записи файла (проверьте права на static/images/banners): %s", e)
            flash('Ошибка сохранения изображения. Проверьте права на папку static/images/banners на сервере.', 'danger')
        except Exception as e:
            logger.exception("Admin homeblock edit: неожиданная ошибка: %s", e)
            db.session.rollback()
            flash(f'Ошибка: {type(e).__name__}. Проверьте логи: journalctl -u lilstore -n 50', 'danger')
    return render_template('admin_homeblock_edit.html', block=block)


@app.route('/admin/homeblock/<int:block_id>/delete', methods=['POST'])
def admin_homeblock_delete(block_id):
    if not _admin_key_valid():
        return "Доступ запрещён", 403
    block = db.get_or_404(HomeBlock, block_id)
    db.session.delete(block)
    db.session.commit()
    _invalidate_cache()
    return redirect(url_for('admin', tab='banners'))


@app.route('/sw.js')
def service_worker():
    """Service Worker для PWA (должен быть в корне для scope /)"""
    return send_from_directory(app.static_folder, 'sw.js', mimetype='application/javascript')


@app.route('/manifest.json')
def manifest():
    """PWA manifest — установка как приложение"""
    base = request.url_root.rstrip('/')
    return jsonify({
        'name': 'LIL STORE',
        'short_name': 'LIL STORE',
        'description': 'IQOS и стики TEREA. Оригинальная продукция, доставка 1–2 дня.',
        'start_url': base + '/',
        'display': 'standalone',
        'background_color': '#f8f9fa',
        'theme_color': '#0d6efd',
        'orientation': 'portrait-primary',
        'icons': [
            {'src': base + url_for('static', filename='LOGO3.png'), 'sizes': '192x192', 'type': 'image/png', 'purpose': 'any'},
            {'src': base + url_for('static', filename='LOGO3.png'), 'sizes': '512x512', 'type': 'image/png', 'purpose': 'any'},
        ],
    })


@app.route('/robots.txt')
def robots():
    """robots.txt для поисковиков"""
    try:
        base = _sitemap_base_url()
    except Exception:
        base = request.url_root.rstrip('/')
    return Response(
        'User-agent: *\n'
        'Allow: /\n'
        'Disallow: /admin\n'
        'Disallow: /banner-click\n'
        'Sitemap: ' + base + url_for('sitemap') + '\n',
        mimetype='text/plain'
    )


def _sitemap_base_url():
    """Базовый URL для sitemap: SITE_URL из config или request (для продакшена — всегда корректный)"""
    try:
        import config
        site_url = getattr(config, 'SITE_URL', None)
        if site_url and isinstance(site_url, str) and site_url.startswith('http'):
            return site_url.rstrip('/')
    except ImportError:
        pass
    return request.url_root.rstrip('/')


@app.route('/sitemap.xml')
def sitemap():
    """Генерация sitemap.xml для поисковиков (кэш 1 ч)"""
    cached = cache.get('sitemap_xml')
    if cached:
        return Response(cached, mimetype='application/xml')
    try:
        base = _sitemap_base_url()
        from datetime import datetime as _dt
        def _lastmod(d):
            return _dt.strftime(d, '%Y-%m-%d') if d else None
        latest = db.session.query(db.func.max(Product.updated_at)).scalar() or db.session.query(db.func.max(Product.created_at)).scalar()
        lastmod_default = _lastmod(latest) if latest else None
        pages = [
            {'loc': base + url_for('index'), 'changefreq': 'daily', 'priority': '1.0', 'lastmod': lastmod_default},
            {'loc': base + url_for('catalog'), 'changefreq': 'daily', 'priority': '0.9', 'lastmod': lastmod_default},
            {'loc': base + url_for('about'), 'changefreq': 'monthly', 'priority': '0.5', 'lastmod': lastmod_default},
            {'loc': base + url_for('contacts'), 'changefreq': 'monthly', 'priority': '0.5', 'lastmod': lastmod_default},
            {'loc': base + url_for('delivery'), 'changefreq': 'monthly', 'priority': '0.5', 'lastmod': lastmod_default},
            {'loc': base + url_for('faq'), 'changefreq': 'monthly', 'priority': '0.6', 'lastmod': lastmod_default},
            {'loc': base + url_for('blog_index'), 'changefreq': 'weekly', 'priority': '0.7', 'lastmod': lastmod_default},
            {'loc': base + url_for('compare'), 'changefreq': 'weekly', 'priority': '0.6', 'lastmod': lastmod_default},
            {'loc': base + url_for('privacy'), 'changefreq': 'yearly', 'priority': '0.3', 'lastmod': lastmod_default},
        ]
        for cat in Category.query.all():
            cat_lastmod = _lastmod(getattr(cat, 'updated_at', None) or cat.created_at)
            pages.append({'loc': base + url_for('catalog', category_slug=cat.slug), 'changefreq': 'weekly', 'priority': '0.8', 'lastmod': cat_lastmod})
        for dm in DeviceModel.query.filter(DeviceModel.slug.isnot(None)).all():
            pages.append({
                'loc': base + url_for('catalog', category_slug=dm.slug),
                'changefreq': 'weekly', 'priority': '0.75', 'lastmod': lastmod_default,
            })
        for post in BlogPost.query.filter_by(is_published=True).all():
            p_lm = _lastmod(getattr(post, 'updated_at', None) or post.created_at)
            pages.append({
                'loc': base + url_for('blog_post', slug=post.slug),
                'changefreq': 'monthly', 'priority': '0.65', 'lastmod': p_lm,
            })
        for p in Product.query.filter_by(in_stock=True).all():
            p_lastmod = _lastmod(getattr(p, 'updated_at', None) or p.created_at)
            pages.append({'loc': base + url_for('product', product_slug=p.get_url_slug()), 'changefreq': 'weekly', 'priority': '0.7', 'lastmod': p_lastmod})
        from xml.sax.saxutils import escape as xml_escape
        xml = ['<?xml version="1.0" encoding="UTF-8"?>', '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">']
        for p in pages:
            lm = '<lastmod>{}</lastmod>'.format(p['lastmod']) if p.get('lastmod') else ''
            xml.append('<url><loc>{}</loc>{}<changefreq>{}</changefreq><priority>{}</priority></url>'.format(
                xml_escape(p['loc']), lm, p['changefreq'], p['priority']))
        xml.append('</urlset>')
        result = '\n'.join(xml)
        cache.set('sitemap_xml', result, timeout=3600)
        return Response(result, mimetype='application/xml')
    except Exception as e:
        logger.exception("Sitemap generation failed: %s", e)
        fallback = (
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
            f'<url><loc>{_sitemap_base_url() + url_for("index")}</loc><changefreq>daily</changefreq><priority>1.0</priority></url>\n'
            f'<url><loc>{_sitemap_base_url() + url_for("catalog")}</loc><changefreq>daily</changefreq><priority>0.9</priority></url>\n'
            '</urlset>'
        )
        return Response(fallback, mimetype='application/xml')


@app.route('/compare')
def compare():
    """Сравнение товаров — товары по id из query ids=1,2,3 (хранятся в localStorage)"""
    ids_str = request.args.get('ids', '')
    ids = []
    for x in ids_str.replace(' ', '').split(','):
        try:
            i = int(x)
            if i > 0:
                ids.append(i)
        except ValueError:
            pass
    ids = list(dict.fromkeys(ids))[:4]  # уникальные, макс 4
    products = Product.query.filter(Product.id.in_(ids)).all() if ids else []
    products_by_id = {p.id: p for p in products}
    products = [products_by_id[i] for i in ids if i in products_by_id]
    return render_template('compare.html', products=products, compare_ids=ids)


@app.route('/favorites')
def favorites():
    """Избранное — товары по id из query ids=1,2,3 (хранятся в localStorage на клиенте)"""
    ids_str = request.args.get('ids', '')
    ids = []
    for x in ids_str.replace(' ', '').split(','):
        try:
            i = int(x)
            if i > 0:
                ids.append(i)
        except ValueError:
            pass
    ids = list(dict.fromkeys(ids))[:50]  # уникальные, макс 50
    products = Product.query.filter(Product.id.in_(ids)).all() if ids else []
    # Сохраняем порядок из ids
    products_by_id = {p.id: p for p in products}
    products = [products_by_id[i] for i in ids if i in products_by_id]
    return render_template('favorites.html', products=products, favorites_ids=ids)


def _escape_like(value):
    """Экранирует % и _ для безопасного использования в LIKE."""
    if not value:
        return value
    return value.replace('\\', '\\\\').replace('%', '\\%').replace('_', '\\_')


@app.route('/search')
def search():
    """Поиск товаров (с синонимами: илюма → iluma и т.д.)"""
    from models import Product
    from search_utils import product_search_filter

    query = (request.args.get('q') or '').strip()
    if query:
        filt = product_search_filter(query, Product)
        products = (
            Product.query.filter(filt).order_by(Product.is_hit.desc(), Product.name).all()
            if filt is not None else []
        )
    else:
        products = []
    return render_template('search.html', products=products, query=query)

def migrate_review_status():
    """Миграция: добавить колонку status к отзывам (для старых БД)"""
    with app.app_context():
        try:
            from sqlalchemy import text
            db.session.execute(text("ALTER TABLE review ADD COLUMN status VARCHAR(20) DEFAULT 'approved'"))
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            if 'duplicate column name' not in str(e).lower():
                print(f"[Migrate] {e}")
        # Обновить старые записи (status IS NULL)
        try:
            for r in Review.query.filter(Review.status == None).all():
                r.status = 'approved'
            db.session.commit()
        except Exception:
            db.session.rollback()


def migrate_telegram_and_cost():
    """Миграция: TelegramUser, Product.cost, BotSetting"""
    with app.app_context():
        from sqlalchemy import text
        # Product.cost
        try:
            db.session.execute(text("ALTER TABLE product ADD COLUMN cost REAL"))
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            if 'duplicate column name' not in str(e).lower():
                print(f"[Migrate] product.cost: {e}")
        db.create_all()


def migrate_product_filters():
    """Миграция: Product.model, color, is_exclusive, is_hit + категория Эксклюзивы"""
    with app.app_context():
        from sqlalchemy import text
        for col, ctype in [('model', 'VARCHAR(50)'), ('color', 'VARCHAR(50)'), ('is_exclusive', 'BOOLEAN'), ('is_hit', 'BOOLEAN')]:
            try:
                db.session.execute(text(f'ALTER TABLE product ADD COLUMN {col} {ctype}'))
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                if 'duplicate column name' not in str(e).lower():
                    print(f"[Migrate] product.{col}: {e}")
        # Категория Эксклюзивы
        cat = Category.query.filter_by(slug='exclusive').first()
        if not cat:
            cat = Category(name='Эксклюзивы', slug='exclusive', description='Лимитированные и эксклюзивные модели')
            db.session.add(cat)
            db.session.commit()
        db.create_all()


def populate_promo_and_hits():
    """Заполняет old_price и is_hit для отображения блоков «Акция!» и «Хит продаж»"""
    with app.app_context():
        products = Product.query.order_by(Product.id).all()
        if not products:
            return
        # Акция: до 6 товаров со скидкой (old_price > price)
        promo_count = Product.query.filter(Product.old_price.isnot(None), Product.old_price > 0).count()
        if promo_count < 6:
            for p in products:
                if promo_count >= 6:
                    break
                if p.old_price is None or p.old_price <= 0:
                    p.old_price = round(p.price * 1.18, 0)  # старая цена на 18% выше
                    promo_count += 1
        # Хит продаж: до 8 товаров (если меньше — дополняем)
        hit_count = Product.query.filter(Product.is_hit == True).count()
        if hit_count < 8:
            for p in products:
                if hit_count >= 8:
                    break
                if not p.is_hit:
                    p.is_hit = True
                    hit_count += 1
        db.session.commit()


def populate_product_colors():
    """Присваивает цвет товарам по названию/описанию (для фильтра каталога). Только товарам без цвета."""
    try:
        from assign_product_colors import detect_color
    except ImportError:
        return
    with app.app_context():
        from sqlalchemy import or_
        products = Product.query.filter(or_(Product.color.is_(None), Product.color == '')).all()
        if not products:
            return
        for p in products:
            color = detect_color(p.name, p.description)
            if color:
                p.color = color
        db.session.commit()


def migrate_banner_badge():
    """Миграция: Banner.badge_type"""
    with app.app_context():
        try:
            from sqlalchemy import text
            db.session.execute(text("ALTER TABLE banner ADD COLUMN badge_type VARCHAR(20)"))
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            if 'duplicate column name' not in str(e).lower():
                print(f"[Migrate] banner.badge_type: {e}")


def migrate_banner_ab_stats():
    """Миграция: Banner.impressions, clicks, ab_test_group для A/B-тестов"""
    with app.app_context():
        from sqlalchemy import text
        for col, ctype in [('impressions', 'INTEGER'), ('clicks', 'INTEGER'), ('ab_test_group', 'VARCHAR(50)')]:
            try:
                db.session.execute(text(f"ALTER TABLE banner ADD COLUMN {col} {ctype}"))
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                if 'duplicate column name' not in str(e).lower():
                    print(f"[Migrate] banner.{col}: {e}")


def migrate_special_banners():
    """Добавить баннеры Telegram, Акция, Скидка, если их ещё нет."""
    with app.app_context():
        if Banner.query.filter(Banner.badge_type == 'telegram').first():
            return  # уже есть
        _tg_url = _get_telegram_bot_url() or '/catalog?on_sale=1'
        _banners_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), app.config['BANNERS_FOLDER'])
        _tg_img = 'telegram_bot.png' if os.path.isfile(os.path.join(_banners_dir, 'telegram_bot.png')) else None
        _prod = Product.query.order_by(Product.id).first()
        _img = _tg_img or (_prod.all_images[0] if _prod and _prod.all_images else None) or (_prod.image if _prod else None) or 'telegram_bot.png'
        # Telegram — картинка из banners/; Акция и Скидка — из products/ если нет
        img_tg = _tg_img or _img or 'telegram_bot.png'
        if img_tg != 'telegram_bot.png' and not img_tg.startswith('products/'):
            img_tg = 'products/' + img_tg
        img_promo = ('products/' + _img) if _img and _img != 'telegram_bot.png' and not _img.startswith('products/') else (_img or 'telegram_bot.png')
        specials = [
            (-3, img_tg, 'ИСПОЛЬЗУЙ TELEGRAM БОТ!',
             '• Актуальные цены и оперативная поддержка\n• Быстро, удобно, круглосуточно',
             'Перейти в бот', _tg_url, 'telegram'),
            (-2, img_promo, 'Акция! Успейте забронировать по специальной цене',
             'Специальные цены на товары', 'Смотреть все акции', '/catalog?on_sale=1', 'promo'),
            (-1, img_promo, 'Скидка 500 ₽ за отзыв',
             'Оставьте отзыв о покупке — получите скидку 500 ₽ на следующий заказ',
             'Написать в ТГ' if (_tg_url and _tg_url != '/catalog?on_sale=1') else 'Подробнее', _tg_url, 'bonus'),
        ]
        for so, img, title, sub, btn, url, badge in specials:
            b = Banner(image=img, title=title, subtitle=sub, button_text=btn, button_url=url, badge_type=badge, sort_order=so, is_active=True)
            db.session.add(b)
        db.session.commit()


def migrate_default_banners():
    """Создать 5 баннеров с товарами (устройства и стики), если баннеров нет. Не трогает существующие."""
    with app.app_context():
        if Banner.query.count() > 0:
            return
        # Устройства (iqos-iluma) и стики (terea-sticks) — до 5 товаров
        cat_devices = Category.query.filter_by(slug='iqos-iluma').first()
        cat_sticks = Category.query.filter_by(slug='terea-sticks').first()
        products_for_banners = []
        if cat_devices:
            products_for_banners.extend(
                Product.query.filter_by(category_id=cat_devices.id).order_by(Product.views.desc(), Product.id).limit(3).all()
            )
        if cat_sticks:
            products_for_banners.extend(
                Product.query.filter_by(category_id=cat_sticks.id).order_by(Product.views.desc(), Product.id).limit(2).all()
            )
        # Добить до 5 из любых товаров
        seen = {p.id for p in products_for_banners}
        if len(products_for_banners) < 5 and seen:
            more = Product.query.filter(~Product.id.in_(seen)).order_by(Product.id).limit(5 - len(products_for_banners)).all()
            products_for_banners.extend(more)
        elif len(products_for_banners) < 5:
            more = Product.query.order_by(Product.id).limit(5).all()
            products_for_banners = more
        else:
            products_for_banners = products_for_banners[:5]
        products_for_banners = products_for_banners[:5]
        if not products_for_banners:
            return
        for i, p in enumerate(products_for_banners):
            img = p.all_images[0] if p.all_images else p.image
            if not img:
                continue
            # Изображения товаров — в products/, сохраняем путь для шаблона
            img_path = 'products/' + img if not img.startswith('products/') else img
            badge = 'hit' if p.is_hit else ('promo' if p.old_price and p.old_price > 0 else None)
            b = Banner(
                image=img_path,
                title=p.name,
                subtitle=f"{p.price:,.0f}".replace(',', ' ') + ' ₽',
                button_text='Купить сейчас',
                button_url=None,
                product_id=p.id,
                badge_type=badge,
                sort_order=i,
                is_active=True,
            )
            db.session.add(b)
        db.session.commit()


def migrate_promo_and_order():
    """Миграция: PromoCode, Order.promo_code, Order.discount_amount, Order.courier_telegram_id"""
    with app.app_context():
        from sqlalchemy import text
        for col, ctype in [('promo_code', 'VARCHAR(50)'), ('discount_amount', 'REAL'), ('courier_telegram_id', 'BIGINT')]:
            try:
                db.session.execute(text(f'ALTER TABLE "order" ADD COLUMN {col} {ctype}'))
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                if 'duplicate column name' not in str(e).lower() and 'no such column' not in str(e).lower():
                    print(f"[Migrate] order.{col}: {e}")
        db.create_all()


if __name__ == '__main__':
    skip_migrations = os.environ.get('SKIP_STARTUP_MIGRATIONS', '').lower() in ('1', 'true', 'yes')
    if skip_migrations:
        logger.info("SKIP_STARTUP_MIGRATIONS=1 — пропуск Alembic и миграций данных")
    if not skip_migrations:
        with app.app_context():
            # Схема БД — через Alembic (вместо ручных ALTER TABLE)
            try:
                from alembic_runner import run_alembic
                run_alembic()
            except Exception as e:
                logger.warning("Alembic: %s (попытка db.create_all)", e)
                db.create_all()
            # Миграции данных (баннеры, цвета, хиты)
            migrate_default_banners()
            migrate_special_banners()
            populate_product_colors()
            if Category.query.count() == 0:
                from create_test_data import create_test_data
                create_test_data()
            populate_promo_and_hits()
    
    debug = os.environ.get('FLASK_DEBUG', 'false').lower() in ('1', 'true', 'yes')
    app.run(debug=debug)